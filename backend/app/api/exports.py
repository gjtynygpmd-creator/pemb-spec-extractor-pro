from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.models.project import ExtractedField, Project

router = APIRouter(prefix="/exports", tags=["exports"])


def _safe_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return cleaned.strip("_") or "project"


def _load(project_id: str, db: Session):
    project = db.get(Project, project_id)
    if not project:
        raise HTTPException(404, "Project not found")
    fields = db.scalars(
        select(ExtractedField)
        .where(ExtractedField.project_id == project_id)
        .order_by(ExtractedField.category, ExtractedField.field_name, ExtractedField.created_at)
    ).all()
    return project, fields


def _field_rows(fields):
    for f in fields:
        yield [
            f.category,
            f.field_name,
            f.value or "",
            f.normalized_value or "",
            f.status,
            round((f.confidence or 0) * 100, 1) if f.confidence is not None else "",
            f.source_file or "",
            f.source_sheet or "",
            f.source_page or "",
            f.source_excerpt or "",
        ]


def _latest_field_map(fields):
    result = {}
    priority = {"accepted": 4, "review": 3, "conflict": 2, "not_applicable": 1}
    for field in fields:
        current = result.get(field.field_name)
        if current is None or priority.get(field.status, 0) >= priority.get(current.status, 0):
            result[field.field_name] = field
    return result


@router.get("/projects/{project_id}/csv")
def export_csv(project_id: str, db: Session = Depends(get_db)):
    project, fields = _load(project_id, db)
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(["Category", "Field", "Value", "Normalized Value", "Review Status", "Confidence %", "Source File", "Source Sheet", "Source Page", "Source Excerpt"])
    writer.writerows(_field_rows(fields))
    content = output.getvalue().encode("utf-8-sig")
    filename = f"{_safe_name(project.name)}_PEMB_Data.csv"
    return StreamingResponse(io.BytesIO(content), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/projects/{project_id}/zoho-csv")
def export_zoho_csv(project_id: str, db: Session = Depends(get_db)):
    project, fields = _load(project_id, db)
    mapping = _latest_field_map(fields)
    standard_columns = [
        "Bid Due", "Customer", "Project Name", "Project Address", "Building Width", "Building Length",
        "Total Square Feet", "Frame Type", "Ridge Offset", "BSW Eave Height", "FSW Eave Height",
        "Roof Panel Type", "Front Roof Slope", "Back Roof Slope", "Roof Panel Gauge", "Wall Panel Type",
        "Wall Panel Gauge", "Roof Insulation", "Wall Insulation", "Roof Insulation Type", "Roof Insulation R-Value",
        "Roof Insulation Thickness", "Roof Insulation Facing", "Wall Insulation Type", "Wall Insulation R-Value",
        "Wall Insulation Thickness", "Wall Insulation Facing", "Risk Category", "Building Code",
        "Roof Live Load", "Dead Load", "Collateral Load", "Ground Snow Load", "Roof Snow Load",
        "Basic Wind Speed", "Wind Exposure", "Site Class", "Seismic Design Category", "S1", "Ss", "Estimator Notes"
    ]
    values = {
        "Project Name": project.name,
        "Customer": project.customer or "",
        "Project Address": project.address or (mapping.get("Project Address").value if mapping.get("Project Address") else ""),
        "Bid Due": project.bid_due.isoformat() if project.bid_due else "",
    }
    for column in standard_columns:
        if column not in values:
            field = mapping.get(column)
            values[column] = field.value if field and field.value else ""
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=standard_columns)
    writer.writeheader()
    writer.writerow(values)
    content = output.getvalue().encode("utf-8-sig")
    filename = f"{_safe_name(project.name)}_Zoho_Import.csv"
    return StreamingResponse(io.BytesIO(content), media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/projects/{project_id}/xlsx")
def export_xlsx(project_id: str, db: Session = Depends(get_db)):
    project, fields = _load(project_id, db)
    wb = Workbook()
    summary = wb.active
    summary.title = "Project Summary"
    summary.append(["PEMB Spec Extractor Pro", "v1.7.2 Export Hotfix"])
    summary.append(["Project Name", project.name])
    summary.append(["Customer", project.customer or ""])
    summary.append(["Address", project.address or ""])
    summary.append(["Bid Due", project.bid_due.isoformat() if project.bid_due else ""])
    summary.append(["Project Status", project.status])
    summary.append(["Exported UTC", datetime.now(timezone.utc).isoformat()])
    summary.append(["Extracted Field Count", len(fields)])
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 70
    summary["A1"].font = Font(bold=True, size=14)
    summary["B1"].font = Font(bold=True)

    data = wb.create_sheet("Extracted Data")
    headers = ["Category", "Field", "Value", "Normalized Value", "Review Status", "Confidence %", "Source File", "Source Sheet", "Source Page", "Source Excerpt"]
    data.append(headers)
    for row in _field_rows(fields):
        data.append(row)
    for cell in data[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F4B183")
        cell.alignment = Alignment(horizontal="center")
    data.freeze_panes = "A2"
    data.auto_filter.ref = data.dimensions
    widths = [20, 30, 32, 24, 16, 14, 36, 14, 12, 70]
    for index, width in enumerate(widths, 1):
        data.column_dimensions[get_column_letter(index)].width = width
    for row in data.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
        row[9].alignment = Alignment(wrap_text=True, vertical="top")

    zoho = wb.create_sheet("Estimator Input")
    mapping = _latest_field_map(fields)
    columns = [
        "Bid Due", "Customer", "Project Name", "Project Address", "Building Width", "Building Length",
        "Total Square Feet", "Frame Type", "Ridge Offset", "BSW Eave Height", "FSW Eave Height",
        "Roof Panel Type", "Front Roof Slope", "Back Roof Slope", "Roof Panel Gauge", "Wall Panel Type",
        "Wall Panel Gauge", "Roof Insulation", "Wall Insulation", "Roof Insulation Type", "Roof Insulation R-Value",
        "Roof Insulation Thickness", "Roof Insulation Facing", "Wall Insulation Type", "Wall Insulation R-Value",
        "Wall Insulation Thickness", "Wall Insulation Facing", "Risk Category", "Building Code",
        "Roof Live Load", "Dead Load", "Collateral Load", "Ground Snow Load", "Roof Snow Load",
        "Basic Wind Speed", "Wind Exposure", "Site Class", "Seismic Design Category", "S1", "Ss", "Estimator Notes"
    ]
    zoho.append(columns)
    row = []
    for column in columns:
        if column == "Project Name": value = project.name
        elif column == "Customer": value = project.customer or ""
        elif column == "Project Address": value = project.address or (mapping.get(column).value if mapping.get(column) else "")
        elif column == "Bid Due": value = project.bid_due.isoformat() if project.bid_due else ""
        else:
            field = mapping.get(column)
            value = field.value if field and field.value else ""
        row.append(value)
    zoho.append(row)
    for cell in zoho[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F4B183")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    zoho.freeze_panes = "A2"
    for index in range(1, len(columns) + 1):
        zoho.column_dimensions[get_column_letter(index)].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{_safe_name(project.name)}_Estimator_Workbook.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


REQUIRED_EXPORT_FIELDS = [
    ("Project", "Project Address"),
    ("Geometry", "Building Width"),
    ("Geometry", "Building Length"),
    ("Geometry", "Building Orientation"),
    ("Framing", "Frame Type"),
    ("Geometry", "BSW Eave Height"),
    ("Geometry", "FSW Eave Height"),
    ("Geometry", "Ridge Offset"),
    ("Geometry", "Front Roof Slope"),
    ("Geometry", "Back Roof Slope"),
    ("Codes & Loads", "Building Code"),
    ("Codes & Loads", "Risk Category"),
    ("Codes & Loads", "Basic Wind Speed"),
    ("Codes & Loads", "Wind Exposure"),
    ("Codes & Loads", "Ground Snow Load"),
    ("Codes & Loads", "Seismic Design Category"),
    ("Envelope", "Roof Panel Type"),
    ("Envelope", "Roof Panel Gauge"),
    ("Envelope", "Wall Panel Type"),
    ("Envelope", "Wall Panel Gauge"),
    ("Insulation", "Roof Insulation Type"),
    ("Insulation", "Roof Insulation R-Value"),
    ("Insulation", "Roof Insulation Thickness"),
    ("Insulation", "Roof Insulation Facing"),
    ("Insulation", "Wall Insulation Type"),
    ("Insulation", "Wall Insulation R-Value"),
    ("Insulation", "Wall Insulation Thickness"),
    ("Insulation", "Wall Insulation Facing"),
    ("Openings", "Overhead Door"),
    ("Accessories", "Gutters"),
    ("Accessories", "Downspouts"),
    ("Finishes", "Paint System"),
]


def _pdf_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _build_pdf(project, fields) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.55 * inch,
        leftMargin=0.55 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.55 * inch,
        title=f"{project.name} - PEMB Estimator Summary",
        author="PEMB Spec Extractor Pro",
    )
    styles = getSampleStyleSheet()
    orange = colors.HexColor("#E66A2C")
    dark = colors.HexColor("#151515")
    light = colors.HexColor("#F3F4F6")
    muted = colors.HexColor("#5F6670")
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=20, leading=23, textColor=dark, alignment=TA_LEFT, spaceAfter=6)
    eyebrow = ParagraphStyle("Eyebrow", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=orange, spaceAfter=4)
    h2 = ParagraphStyle("H2Custom", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=12, leading=14, textColor=dark, spaceBefore=8, spaceAfter=6)
    body = ParagraphStyle("BodyCustom", parent=styles["BodyText"], fontSize=8.5, leading=11, textColor=dark)
    small = ParagraphStyle("Small", parent=body, fontSize=7.5, leading=9, textColor=muted)
    white_body = ParagraphStyle("WhiteBody", parent=body, textColor=colors.white, fontName="Helvetica-Bold")
    story = [
        Paragraph("PEMB SPEC EXTRACTOR PRO - v1.7.2 EXPORT HOTFIX", eyebrow),
        Paragraph(_pdf_text(project.name), title_style),
        Paragraph("Estimator summary generated from reviewed, extracted, and manually entered project data.", small),
        Spacer(1, 10),
    ]
    mapping = _latest_field_map(fields)
    project_rows = [
        ["Project", _pdf_text(project.name), "Status", _pdf_text(project.status)],
        ["Customer", _pdf_text(project.customer or "Not provided"), "Bid Due", _pdf_text(project.bid_due.isoformat() if project.bid_due else "Not provided")],
        ["Address", _pdf_text((project.address or ((mapping.get("Project Address").value if mapping.get("Project Address") else None)) or "Not provided")), "Exported UTC", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
    ]
    project_table = Table([[Paragraph(str(c), body) for c in row] for row in project_rows], colWidths=[0.85*inch, 2.45*inch, 0.85*inch, 2.45*inch])
    project_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (0,-1), light), ("BACKGROUND", (2,0), (2,-1), light),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"), ("FONTNAME", (2,0), (2,-1), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.35, colors.HexColor("#D7DBE0")),
        ("VALIGN", (0,0), (-1,-1), "TOP"), ("LEFTPADDING", (0,0), (-1,-1), 6), ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 5), ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))
    story.extend([project_table, Spacer(1, 10)])

    accepted = sum(1 for f in fields if f.status == "accepted")
    conflicts = sum(1 for f in fields if f.status == "conflict")
    missing = [(cat, name) for cat, name in REQUIRED_EXPORT_FIELDS if not mapping.get(name) or not (mapping[name].value or "").strip()]
    metrics = [["Captured Fields", str(len(fields)), "Accepted", str(accepted), "Conflicts", str(conflicts), "Missing Core", str(len(missing))]]
    metrics_table = Table([[Paragraph(c, white_body) for c in metrics[0]]], colWidths=[0.9*inch,0.45*inch,0.65*inch,0.45*inch,0.65*inch,0.45*inch,0.75*inch,0.45*inch])
    metrics_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), dark), ("TEXTCOLOR", (0,0), (-1,-1), colors.white),
        ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"), ("ALIGN", (1,0), (1,0), "CENTER"),
        ("ALIGN", (3,0), (3,0), "CENTER"), ("ALIGN", (5,0), (5,0), "CENTER"), ("ALIGN", (7,0), (7,0), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"), ("TOPPADDING", (0,0), (-1,-1), 6), ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    story.extend([metrics_table, Spacer(1, 10), Paragraph("Captured Estimating Data", h2)])

    if fields:
        grouped = {}
        for field in fields:
            grouped.setdefault(field.category or "Other", []).append(field)
        first_group = True
        for category, items in grouped.items():
            if not first_group and len(story) > 20:
                story.append(Spacer(1, 4))
            first_group = False
            story.append(Paragraph(_pdf_text(category), ParagraphStyle("Category", parent=h2, fontSize=10, leading=12, textColor=orange, spaceBefore=5, spaceAfter=3)))
            rows = [[Paragraph("Field", body), Paragraph("Value", body), Paragraph("Status", body), Paragraph("Source", body)]]
            for f in items:
                source = " / ".join([x for x in [f.source_file, f.source_sheet, f"p. {f.source_page}" if f.source_page else None] if x]) or "Manual / not recorded"
                rows.append([Paragraph(_pdf_text(f.field_name), body), Paragraph(_pdf_text(f.value or "Not provided"), body), Paragraph(_pdf_text(f.status), small), Paragraph(_pdf_text(source), small)])
            table = Table(rows, colWidths=[1.55*inch,2.35*inch,0.75*inch,2.15*inch], repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), orange), ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#D7DBE0")),
                ("VALIGN", (0,0), (-1,-1), "TOP"), ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, light]),
                ("LEFTPADDING", (0,0), (-1,-1), 5), ("RIGHTPADDING", (0,0), (-1,-1), 5),
                ("TOPPADDING", (0,0), (-1,-1), 4), ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ]))
            story.append(table)
    else:
        story.append(Paragraph("No estimating fields have been captured yet. Use the Missing Information list in the project workspace to enter confirmed values, then export the PDF again.", body))

    story.extend([Spacer(1, 10), Paragraph("Missing Core Information", h2)])
    if missing:
        missing_rows = [[Paragraph("Category", white_body), Paragraph("Required Field", white_body)]] + [[Paragraph(_pdf_text(cat), body), Paragraph(_pdf_text(name), body)] for cat, name in missing]
        missing_table = Table(missing_rows, colWidths=[1.6*inch,5.2*inch], repeatRows=1)
        missing_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), dark), ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"), ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#D7DBE0")),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, light]), ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("LEFTPADDING", (0,0), (-1,-1), 5), ("RIGHTPADDING", (0,0), (-1,-1), 5),
            ("TOPPADDING", (0,0), (-1,-1), 4), ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))
        story.append(missing_table)
    else:
        story.append(Paragraph("All core estimating fields currently have values.", body))

    story.extend([Spacer(1, 12), Paragraph("Estimator note: Confirm all extracted values against the governing drawings, specifications, addenda, and applicable building code before pricing or design.", small)])

    def page_footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(orange)
        canvas.setLineWidth(1.2)
        canvas.line(doc.leftMargin, 0.42*inch, letter[0]-doc.rightMargin, 0.42*inch)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(muted)
        canvas.drawString(doc.leftMargin, 0.27*inch, "PEMB Spec Extractor Pro - Estimator Summary")
        canvas.drawRightString(letter[0]-doc.rightMargin, 0.27*inch, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=page_footer, onLaterPages=page_footer)
    buffer.seek(0)
    return buffer.getvalue()


@router.get("/projects/{project_id}/pdf")
def export_pdf(project_id: str, db: Session = Depends(get_db)):
    project, fields = _load(project_id, db)
    content = _build_pdf(project, fields)
    filename = f"{_safe_name(project.name)}_Estimator_Summary.pdf"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
